import openpyxl
from config.settings import BASE_DIR

def genExcel(transacciones):

    doc = openpyxl.load_workbook(BASE_DIR + '/loptique/static/' + 'sheet.xlsx')

    sheets = doc.sheetnames #.sheetnames accede a las hojas dentro del doc y las convierte en lista

    hoja1 = doc['Hoja 1']#Guardamos en una variable la info de una hoja

    hoja1['A3'] = "Fecha"
    hoja1['B3'] = "Descripcion"
    hoja1['C3'] = "Ingreso"
    hoja1['D3'] = "Egreso"

    total = 0
    for i in range(len(transacciones)):
        minutos = str(transacciones[i].fecha.minute)
        if len(minutos) == 1:
            minutos = "0" + str(transacciones[i].fecha.minute)
        hora = transacciones[i].fecha.hour - 3
        if hora < 0:
            hora = 24 + hora
        hora = str(hora)
        if len(hora) == 1:
            hora = "0" + hora

        hoja1['A' + str(i+4)] = "{}/{}/{} {}:{}".format(str(transacciones[i].fecha.day), str(transacciones[i].fecha.month), str(transacciones[i].fecha.year),  hora,  minutos)
        hoja1['B' + str(i+4)] = transacciones[i].descripcion
        print(transacciones[i].egreso)
        if transacciones[i].egreso:
            hoja1['D' + str(i + 4)] = transacciones[i].monto
        else:
            hoja1['C' + str(i + 4)] = transacciones[i].monto


    doc.save(BASE_DIR + '/loptique/static/' + 'Balance.xlsx')